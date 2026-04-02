from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO


_header_font = Font(bold=True, color="FFFFFF", size=11)
_header_fill = PatternFill(fill_type="solid", fgColor="4F81BD")
_header_align = Alignment(horizontal="center")
_totals_font = Font(bold=True, color="FFFFFF", size=11)
_totals_fill = PatternFill(fill_type="solid", fgColor="4F81BD")
_thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws, col_count: int):
    for cell in ws[1]:
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = _header_align
        cell.border = _thin_border


def _style_data_rows(ws, start_row: int, end_row: int, col_count: int):
    alt_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill


def _auto_column_width(ws, col_count: int):
    for col_idx in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 3


def generate_sales_excel(data: list[dict], filters: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    if data:
        headers = list(data[0].keys())
        col_count = len(headers)
        ws.append(headers)
        for row in data:
            ws.append([row.get(h) for h in headers])

        # Totals row
        total_amount = sum(row.get("Subtotal", 0) for row in data)
        totals_row = ["" for _ in headers]
        totals_row[0] = "TOTAL"
        if "Subtotal" in headers:
            totals_row[headers.index("Subtotal")] = round(total_amount, 2)
        ws.append(totals_row)

        _style_header_row(ws, col_count)
        _style_data_rows(ws, 2, len(data) + 1, col_count)

        # Style totals row
        totals_row_idx = len(data) + 2
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=totals_row_idx, column=col_idx)
            cell.font = _totals_font
            cell.fill = _totals_fill
            cell.border = _thin_border

        _auto_column_width(ws, col_count)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_inventory_excel(data: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Report"

    if data:
        headers = list(data[0].keys())
        col_count = len(headers)
        ws.append(headers)
        for row in data:
            ws.append([row.get(h) for h in headers])

        # Totals row
        total_value = sum(row.get("Total Value", 0) for row in data)
        total_qty = sum(row.get("Quantity", 0) for row in data)
        totals_row = ["" for _ in headers]
        totals_row[0] = "TOTAL"
        if "Quantity" in headers:
            totals_row[headers.index("Quantity")] = total_qty
        if "Total Value" in headers:
            totals_row[headers.index("Total Value")] = round(total_value, 2)
        ws.append(totals_row)

        _style_header_row(ws, col_count)
        _style_data_rows(ws, 2, len(data) + 1, col_count)

        # Style totals row
        totals_row_idx = len(data) + 2
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=totals_row_idx, column=col_idx)
            cell.font = _totals_font
            cell.fill = _totals_fill
            cell.border = _thin_border

        _auto_column_width(ws, col_count)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
